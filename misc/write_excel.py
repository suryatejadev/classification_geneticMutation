from openpyxl import load_workbook
import pickle

path = 'output.xlsx'
wb = load_workbook(path)
ws = wb.get_active_sheet()

a = pickle.load(open('output/results_full.pickle','rb'))
a1 = pickle.load(open('output/results_full_ann.pickle','rb'))
a2 = pickle.load(open('output/results_full_time.pickle','rb'))
a3 = pickle.load(open('output/results_full_dataexp1.pickle','rb'))
print a.shape,a1.shape,a2.shape,a3.shape

index = ['B','C','D','E','F']
# print no gene results
for i in range(5):
    for j in range(5):
        ws[index[j]+str(i+2)] = a3[i,j,0,0]
        ws[index[j]+str(i+2+6)] = a3[i,j,0,1]

wb.save(path)

